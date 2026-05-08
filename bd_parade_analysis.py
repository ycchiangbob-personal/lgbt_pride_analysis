#!/usr/bin/env python3
"""
遊行BD機會分析 — Cross-reference parade participants with commercial sponsorship database.
Usage: python bd_parade_analysis.py
"""

import json
import re
import unicodedata
from dataclasses import dataclass, field
from datetime import date
from pathlib import Path
from typing import Optional

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font
    HAS_EXCEL = True
except ImportError:
    HAS_EXCEL = False

# ─── Layer 0: Constants & Utilities ──────────────────────────────────────────

RECENCY_WEIGHTS = {
    '2025': 1.0, '2024': 0.9, '2023': 0.8,
    '2022': 0.7, '2020': 0.5, '2019': 0.4,
}
SCORED_YEARS = set(RECENCY_WEIGHTS.keys())
FLOAT_TIERS = {'花車', '車', '車（頭車）'}

CIVIC_SIGNALS = (
    '社團法人', '財團法人', '協會', '學會', '聯盟', '基金會',
    '大學', '學院', '研究所', '研究社', '黨部', '黨',
    '陣線', '連線', '工作坊', '促進會', '部屋',
    '童軍', '教會', '委員會', '廢核',
    '系學會', '所學會', '讀書會', '課程',
    '社大', '非營利', '公民行動',
    '親子共學', '學生聯合', '學生議會', '學生會', '系學會',
    # Named orgs that lack standard civic suffixes
    '時代力量', '台灣基進', '綠黨', '台灣綠黨', '社會民主黨',
    '台灣民眾黨', '自由台灣黨', '平權大平台', '彩虹平權大平台', '彩虹平權',
    'Out in Taiwan', '彩虹前線', '樂生保留', '青年樂生',
    # Performing arts / community social groups (not commercial)
    '管樂團', '合唱團', '樂團契', '吉他社', '游泳社', '跑步',
    'Asians and Friends', 'Frontrunners',
    '跨國同婚', '多元教育', '同志父母', '勵馨',
    # Civic / diaspora / community groups without standard suffixes
    '協進會', '同志會', '邊城青年', '性壇社', '大馬旅台',
    '多元性別社群', '社群協進', 'mamak',
    # Identity / advocacy groups not formally incorporated
    '非二元', '酷兒浪子', '助人工作者', '無浪漫', '無性戀',
    '笨瓜秀', '揪喜樂',
    # Student clubs (university orgs not caught by "大學")
    '男同性戀社', '女同誌異社', '同誌社', '浪達社', 'bdsm社',
    '愛慾實務社', '性別友善社', '女性主義研究社',
    # Outdoor / hiking community groups
    '山棧花', '登山社', '向山社',
    # BDSM community groups, other city pride orgs (not commercial)
    '皮繩愉虐邦', '台南彩虹遊行', '高雄同志大遊行', '花東彩虹',
    '縛.生', '縛生', '桃園彩虹野餐日',
)

# Map normalized parade md name → donor canonical name (id_state.json key)
KNOWN_ALIASES = {
    # AstraZeneca
    'astrazeneca': 'AZ',
    'az': 'AZ',
    '臺灣阿斯特捷利康': 'AZ',
    '台灣阿斯特捷利康': 'AZ',
    '臺灣阿斯特捷利康股份有限公司': 'AZ',
    'astrazeneca（az）': 'AZ',
    # GSK
    'gsk taiwan': 'GSK',
    'gsk 葛蘭素史克': 'GSK',
    '葛蘭素史克': 'GSK',
    # GILEAD
    '吉立亞醫藥': 'GILEAD',
    'gilead sciences 吉立亞醫藥': 'GILEAD',
    'gilead吉立亞醫藥': 'GILEAD',
    'gilead sciences': 'GILEAD',
    # GSTAR (gay bar)
    'g*star': 'GSTAR',
    'g.star': 'GSTAR',
    'gstar': 'GSTAR',
    # LesPark
    'lespark': '拉拉公園',
    'les park': '拉拉公園',
    'lespark拉拉公園': '拉拉公園',
    # Uber
    'uber': 'UBer',
    'uber｜uber eats': 'UBer',
    'uber | uber eats': 'UBer',
    'uberｅats': 'UBer',
    # Standard Chartered
    '渣打國際商業銀行': '渣打銀行',
    # 諾億 variants
    '磊山保經諾億團隊': '諾億',
    '諾億財富管理顧問': '諾億',
    '磊山保經 諾億團隊': '諾億',
    '諾億保險經紀人': '諾億保經',
    # 法國巴黎人壽
    '法商法國巴黎人壽保險股份有限公司台灣分公司': '法國巴黎人壽',
    # Microsoft
    '台灣微軟股份有限公司': 'Microsoft',
    # Google ERGs
    'pride@google taiwan': 'Google',
    'google gaygler': 'Google',
    'pride at google taiwan': 'Google',
    # TSMC ERG
    'pride@tsmc': 'TSMC',
    # GAP
    'gap taiwan': 'GAP',
    '台灣蓋璞': 'GAP',
    # Novo Nordisk
    '台灣諾和諾德': '諾和諾德',
    '諾和諾德藥品股份有限公司（novo nordisk）': '諾和諾德',
    '台灣諾和諾德藥品股份有限公司（novo nordisk）': '諾和諾德',
    '台灣諾和諾德藥品股份有限公司': '諾和諾德',
    # Qualcomm
    '美商高通': '高通',
    'qualcomm 高通半導體': '高通',
    'qualcomm taiwan': '高通',
    # Unilever
    '聯合利華': 'Unilever',
    '聯合利華 為愛站出來 stand by u': 'Unilever',
    # EY
    '安永聯合會計師事務所': 'EY',
    '安永聯合會計師事務所（ey taiwan）': 'EY',
    # Boston Scientific
    '波士頓科技 boston scientific': '波士頓',
    '荷蘭商波士頓科技有限公司台灣分公司': '波士頓',
    'boston scientific taiwan 波士頓科技': '波士頓',
    '荷蘭商瑞力登國際有限公司台灣分公司': '波士頓',
    # MSD / 默沙東
    'msd taiwan': '默沙東',
    '美商默沙東藥廠': '默沙東',
    'msd 美商默沙東藥廠': '默沙東',
    # Micron
    '美光台灣': '美光',
    '美光科技 micron technology': '美光',
    # Citibank
    '花旗台灣商業銀行': '花旗商銀',
    '花旗（台灣）商業銀行': '花旗商銀',
    '花旗(台灣)商業銀行': '花旗商銀',
    # 台虎
    '臺虎精釀': '台虎',
    # Roche
    '羅氏大藥廠與羅氏醫療診斷設備': '羅氏',
    '羅氏大藥廠股份有限公司': '羅氏',
    # Sanofi
    '賽諾菲股份有限公司 sanofi': 'Sanofi',
    '賽諾菲股份有限公司': 'Sanofi',
    # JLR
    'jlr taiwan 台灣捷豹路虎': '捷豹路虎',
    # Porsche
    '台灣保時捷車業股份有限公司': '台灣保時捷車業股份有限公司 \nPorsche Taiwan Motors Limited',
    'porsche taiwan': '台灣保時捷車業股份有限公司 \nPorsche Taiwan Motors Limited',
    '台灣保時捷車業股份有限公司 porsche taiwan motors limited': '台灣保時捷車業股份有限公司 \nPorsche Taiwan Motors Limited',
    # Synopsys
    'synopsys taiwan 台灣新思科技': '新思科技',
    # Schroders
    '施羅德投資': 'Schroders',
    '施羅德證券投資信託股份有限公司（施羅德投信）': 'Schroders',
    '施羅德投信': 'Schroders',
    # MAC = M.A.C (confirmed same company; canonical = M.A.C D0123)
    'mac': 'M.A.C',
    'mac taiwan': 'M.A.C',
    'm.a.c cosmetics taiwan': 'M.A.C',
    'm·a·c cosmetics taiwan': 'M.A.C',
    'm·a·c': 'M.A.C',
    # 台灣保樂力加
    '台灣保樂力加 pernod ricard taiwan': 'Pernod Ricard Taiwan',
    # Fusion
    'fusion 聚變': 'Fusion 聚變',
    # Activision
    '動視暴雪 activision blizzard': 'Activision',
    'activision blizzard（taiwan）': 'Activision',
    'activision blizzard': 'Activision',
    # Dell
    'dell technology': 'Dell',
    'dell technologies': 'Dell',
    'dell technologies | taiwan design center 20 anniversary': 'Dell',
    # Goldman Sachs
    '美商高盛': 'Goldman Sachs',
    # 犀牛盾
    'rhinoshield': '犀牛盾',
    '犀牛盾 rhinoshield': '犀牛盾',
    # Tapestry
    'tapestry': 'Tapestry',
    # VERVE
    'verve': 'VERVE',
    # NIKE
    'nike': 'NIKE',
    # Flying V (星宇航空)
    '星宇航空 starlux airlines': 'Starlux',
    # ── Confirmed by user CSV review (2026-05-03) ──────────────────────────
    # Follow prospects
    'gap（台灣蓋璞）': 'GAP',
    'maryjane pizza': '瑪莉珍比薩',
    '50% fifty percent': 'FiftyPercent',
    'modeltv 影音串流平台': 'Model TV',
    'modeltv': 'Model TV',
    '臺北市觀光傳播局': '台北市觀傳局',
    'gagaoolala lgbtq+ & bl線上影音平台': 'GaGaoLaLa',
    'gagaoolala 屬於你的故事': 'GaGaoLaLa',
    'gagaoolala': 'GaGaoLaLa',
    'gagaoolala lgbtq+': 'GaGaoLaLa',
    '飛比 feebee': '飛比',
    'feebee': '飛比',
    '臺灣阿斯特捷利康（astrazeneca）': 'AZ',
    '金賓美國波本威士忌': '美商百富門',  # Jim Beam = Brown-Forman
    '金賓': '美商百富門',
    '美國高通公司': '高通',
    'pizza hut': '必勝客',
    # Donor ID confirmed by user
    '台灣美光': '美光',                             # D0084
    '台灣樂天集團': '樂天',                           # D0093
    'locker room & fairy taipei': 'Club Locker Room',  # D0219
    'fairy and locker room': 'Club Locker Room',       # D0219
    # Market booth (市集) donors: no canonical in JSON; group variants together
    'lezs 女人國': '女人國',
    "lez's meeting 女人國": '女人國',
    'lezs 女人國/頤創藝': '女人國',
    '女人國': '女人國',
    # Tesla
    'tesla taiwan': 'Tesla Taiwan',
    '台灣特斯拉': 'Tesla Taiwan',
    'tesla 台灣特斯拉': 'Tesla Taiwan',
}


# Canonical names confirmed by user as Follow-up targets (from CSV review 2026-05-03)
USER_FOLLOW_CANONICALS = {
    'GAP', 'FiftyPercent', 'Model TV', '台北市觀傳局', 'GaGaoLaLa',
    '飛比', 'AZ', '美商百富門', '高通', '必勝客', '瑪莉珍比薩',
}


def normalize_name(s: str) -> str:
    if not s:
        return ''
    s = unicodedata.normalize('NFKC', s)
    s = re.sub(r'\s+', ' ', s).strip().lower()
    return s


def smart_split(line: str) -> list:
    """Split by 、 but respect （...） brackets."""
    parts, current, depth = [], [], 0
    for ch in line:
        if ch in '（(':
            depth += 1
            current.append(ch)
        elif ch in '）)':
            depth -= 1
            current.append(ch)
        elif ch == '、' and depth == 0:
            tok = ''.join(current).strip()
            if tok:
                parts.append(tok)
            current = []
        else:
            current.append(ch)
    tok = ''.join(current).strip()
    if tok:
        parts.append(tok)
    return parts


def is_civic(name: str) -> bool:
    return any(sig in name for sig in CIVIC_SIGNALS)


# ─── Layer 1: MD Parser ───────────────────────────────────────────────────────

def parse_md(filepath: Path) -> dict:
    """Returns {year: {'commercial_floats': set, 'parade_teams': set, '2021_all': set|None}}"""
    text = filepath.read_text(encoding='utf-8')
    chunks = re.split(r'(?=## \d{4} 年)', text)
    result = {}
    for chunk in chunks:
        m = re.match(r'## (\d{4}) 年', chunk)
        if not m:
            continue
        year = m.group(1)
        floats, teams = set(), set()
        for line in chunk.splitlines():
            line = line.strip()
            if line.startswith('**商業車：**'):
                content = line[len('**商業車：**'):].strip()
                for name in smart_split(content):
                    n = name.strip()
                    if n:
                        floats.add(n)
            elif line.startswith('**隊伍：**'):
                content = line[len('**隊伍：**'):].strip()
                for name in smart_split(content):
                    n = name.strip()
                    if n:
                        teams.add(n)
        if year == '2021':
            all_2021 = set()
            for line in chunk.splitlines():
                line = line.strip()
                if '**參與團體' in line:
                    # Remove bold markers and leading label
                    content = re.sub(r'\*\*[^*]*\*\*', '', line).strip().lstrip('：').strip()
                    for name in smart_split(content):
                        n = name.strip()
                        if n:
                            all_2021.add(n)
            result[year] = {'commercial_floats': set(), 'parade_teams': set(), '2021_all': all_2021}
        else:
            result[year] = {'commercial_floats': floats, 'parade_teams': teams, '2021_all': None}
    return result


# ─── Layer 2: Donor DB Loader ────────────────────────────────────────────────

@dataclass
class DonorDB:
    canonical_to_id: dict = field(default_factory=dict)
    id_to_canonical: dict = field(default_factory=dict)
    norm_to_canonical: dict = field(default_factory=dict)
    float_buyers: dict = field(default_factory=dict)    # year -> set(canonical)
    tier_sponsors: dict = field(default_factory=dict)   # year -> set(canonical)
    market_booth: dict = field(default_factory=dict)    # year -> set(canonical)
    other_sponsors: dict = field(default_factory=dict)  # year -> set(canonical)


def _classify(tier_orig: str) -> str:
    if not tier_orig:
        return 'other'
    if tier_orig in FLOAT_TIERS:
        return 'float'
    if tier_orig.endswith('級'):
        return 'tier'
    if '市集' in tier_orig:
        return 'market'
    return 'other'


def _add(bucket: dict, year: str, canonical: str):
    bucket.setdefault(year, set()).add(canonical)


def load_donors(base: Path) -> DonorDB:
    db = DonorDB()

    with open(base / 'id_state.json') as f:
        id_data = json.load(f)
    for name, did in id_data['all_donors'].items():
        db.canonical_to_id[name] = did
        db.id_to_canonical[did] = name
        db.norm_to_canonical[normalize_name(name)] = name

    # 2022–2025
    with open(base / 'all_extracted_v2.json') as f:
        extracted = json.load(f)
    for year, records in extracted.items():
        for r in records:
            c = r.get('name_canonical', '')
            cat = _classify(r.get('tier_orig', ''))
            if cat == 'float':
                _add(db.float_buyers, year, c)
            elif cat == 'tier':
                _add(db.tier_sponsors, year, c)
            elif cat == 'market':
                _add(db.market_booth, year, c)
            else:
                _add(db.other_sponsors, year, c)

    # 2020
    with open(base / '2020_classified.json') as f:
        d2020 = json.load(f)
    for r in d2020['records']:
        c = r.get('sponsor_canonical', '')
        cat = _classify(r.get('tier_orig', ''))
        if cat == 'float':
            _add(db.float_buyers, '2020', c)
        elif cat == 'tier':
            _add(db.tier_sponsors, '2020', c)
        else:
            _add(db.other_sponsors, '2020', c)

    # 2019 (tier only in JSON; floats come from md)
    with open(base / '2019_classified.json') as f:
        d2019 = json.load(f)
    for r in d2019['records']:
        c = r.get('sponsor_canonical', '')
        cat = _classify(r.get('tier_orig', ''))
        if cat == 'tier':
            _add(db.tier_sponsors, '2019', c)
        else:
            _add(db.other_sponsors, '2019', c)

    # 2021
    with open(base / '2021_final.json') as f:
        d2021 = json.load(f)
    for tier_key, entries in d2021.items():
        for entry in entries:
            c = entry[1] if len(entry) > 1 else entry[0]
            if c:
                _add(db.tier_sponsors, '2021', c)

    return db


# ─── Layer 3: Name Matcher ───────────────────────────────────────────────────

@dataclass
class MatchResult:
    raw_name: str
    canonical: Optional[str]
    donor_id: Optional[str]
    confidence: str   # exact | alias | substring | unmatched
    note: str = ''


class NameMatcher:
    def __init__(self, db: DonorDB):
        self.db = db
        self.pending: list = []
        self._pending_seen: set = set()
        # normalized alias → canonical name
        self._alias: dict = {normalize_name(k): v for k, v in KNOWN_ALIASES.items()}

    def match(self, raw_name: str) -> MatchResult:
        n = normalize_name(raw_name)

        # 1. Exact
        if n in self.db.norm_to_canonical:
            canonical = self.db.norm_to_canonical[n]
            return MatchResult(raw_name, canonical, self.db.canonical_to_id.get(canonical), 'exact')

        # 2. Alias
        if n in self._alias:
            canonical = self._alias[n]
            did = self.db.canonical_to_id.get(canonical)
            return MatchResult(raw_name, canonical, did, 'alias')

        # 3. Substring (only canonical names ≥ 4 chars; pick longest match)
        best_len, best_c = 0, None
        for norm_c, canonical in self.db.norm_to_canonical.items():
            if len(norm_c) >= 4 and (norm_c in n or n in norm_c):
                if len(norm_c) > best_len:
                    best_len, best_c = len(norm_c), canonical
        if best_c:
            return MatchResult(raw_name, best_c, self.db.canonical_to_id.get(best_c),
                               'substring', f'substring→{best_c}')

        # 4. Unmatched
        if n not in self._pending_seen:
            self._pending_seen.add(n)
            self.pending.append({'raw_name': raw_name, 'normalized': n})
        return MatchResult(raw_name, None, None, 'unmatched')


# ─── Layer 4a: Analysis A — Historical Conversion ────────────────────────────

def analyze_A(md_data: dict, db: DonorDB, matcher: NameMatcher) -> dict:
    """For each commercial float buyer, check if they first appeared as a parade team."""

    # float_entries: identity_key -> {canonical, donor_id, years}
    float_entries: dict = {}

    def float_key(mr: MatchResult, raw: str) -> str:
        return mr.donor_id or (mr.canonical and f'c:{mr.canonical}') or f'r:{normalize_name(raw)}'

    for year in sorted(SCORED_YEARS):
        for raw in md_data.get(year, {}).get('commercial_floats', set()):
            mr = matcher.match(raw)
            k = float_key(mr, raw)
            if k not in float_entries:
                float_entries[k] = {'canonical': mr.canonical or raw, 'donor_id': mr.donor_id, 'years': set()}
            float_entries[k]['years'].add(year)

    # Also add float buyers from JSON db (covers 2020 + 2022–2025)
    for year, canonicals in db.float_buyers.items():
        for canonical in canonicals:
            did = db.canonical_to_id.get(canonical)
            k = did or f'c:{canonical}'
            if k not in float_entries:
                float_entries[k] = {'canonical': canonical, 'donor_id': did, 'years': set()}
            float_entries[k]['years'].add(year)

    # team appearances: identity_key -> set of years
    team_years: dict = {}
    for year in sorted(SCORED_YEARS):
        for raw in md_data.get(year, {}).get('parade_teams', set()):
            mr = matcher.match(raw)
            k = float_key(mr, raw)
            team_years.setdefault(k, set()).add(year)

    detail_rows = []
    had_team_first = 0

    for k, info in float_entries.items():
        first_float = min(info['years'])
        prior = sorted(y for y in team_years.get(k, set()) if y < first_float)
        same = first_float in team_years.get(k, set())
        was_first = bool(prior)
        if was_first:
            had_team_first += 1
        detail_rows.append({
            'name': info['canonical'],
            'donor_id': info['donor_id'] or '',
            'float_years': sorted(info['years']),
            'first_float_year': first_float,
            'team_years_before_float': prior,
            'same_year_both': same,
            'was_team_first': was_first,
        })

    detail_rows.sort(key=lambda r: (r['first_float_year'], not r['was_team_first']))
    total = len(float_entries)

    return {
        'total_float_buyers': total,
        'had_team_history_before_float': had_team_first,
        'conversion_rate_pct': round(had_team_first / total * 100, 1) if total else 0,
        'detail_rows': detail_rows,
    }


# ─── Layer 4b: Analysis B — Prospect Scoring ─────────────────────────────────

def analyze_B(md_data: dict, db: DonorDB, matcher: NameMatcher) -> list:

    def entity_key(mr: MatchResult, raw: str) -> str:
        return mr.donor_id or (mr.canonical and f'c:{mr.canonical}') or f'r:{normalize_name(raw)}'

    # Build entity registry from parade teams
    entities: dict = {}
    for year in sorted(SCORED_YEARS):
        for raw in md_data.get(year, {}).get('parade_teams', set()):
            mr = matcher.match(raw)
            k = entity_key(mr, raw)
            if k not in entities:
                entities[k] = {
                    'canonical': mr.canonical,
                    'donor_id': mr.donor_id,
                    'raw_names': set(),
                    'years': set(),
                    'best_conf': mr.confidence,
                }
            entities[k]['raw_names'].add(raw)
            entities[k]['years'].add(year)
            if mr.confidence in ('exact', 'alias') and entities[k]['best_conf'] not in ('exact', 'alias'):
                entities[k]['canonical'] = mr.canonical
                entities[k]['donor_id'] = mr.donor_id
                entities[k]['best_conf'] = mr.confidence

    # Build float buyer key set (from md + db)
    float_keys: set = set()
    for year in SCORED_YEARS:
        for raw in md_data.get(year, {}).get('commercial_floats', set()):
            mr = matcher.match(raw)
            float_keys.add(entity_key(mr, raw))
    for year, canonicals in db.float_buyers.items():
        for c in canonicals:
            did = db.canonical_to_id.get(c)
            float_keys.add(did or f'c:{c}')

    results = []
    for k, info in entities.items():
        years_sorted = sorted(info['years'])
        score = round(sum(RECENCY_WEIGHTS[y] for y in years_sorted), 2)
        is_float = k in float_keys
        canonical = info['canonical']

        # Donor status (check most recent years first)
        donor_status = '非捐款者'
        for yr in sorted(years_sorted, reverse=True):
            if canonical in db.tier_sponsors.get(yr, set()):
                donor_status = '級別贊助'
                break
            if canonical in db.market_booth.get(yr, set()):
                donor_status = '市集'
                break
            if canonical in db.other_sponsors.get(yr, set()):
                donor_status = '其他'
                break
        if donor_status == '非捐款者' and info['donor_id']:
            # Has ID, check all years
            for yr_set in db.tier_sponsors.values():
                if canonical in yr_set:
                    donor_status = '級別贊助'
                    break
            if donor_status == '非捐款者':
                for yr_set in db.market_booth.values():
                    if canonical in yr_set:
                        donor_status = '市集'
                        break

        # Entity type classification
        primary = canonical or min(info['raw_names'], key=len)
        all_names = ' '.join(info['raw_names'])
        entity_type = '可能商業'
        if is_civic(all_names):
            entity_type = 'NGO/學生/公民社會'
        # ERG: only flag as uncertain if unresolved (no donor_id).
        # If alias already resolved it to a parent company (has donor_id), treat as commercial.
        if any(re.search(r'pride@', n, re.I) for n in info['raw_names']) and not info['donor_id']:
            entity_type = '待確認 (ERG未解析)'
        if any(x in all_names for x in ('辦事處', '貿易辦事處', '文化院', 'AIT', 'European Economic')):
            entity_type = '待確認 (政府機構)'

        is_follow = canonical in USER_FOLLOW_CANONICALS if canonical else False

        results.append({
            'key': k,
            'name_primary': primary,
            'name_variants': sorted(info['raw_names']),
            'canonical': canonical,
            'donor_id': info['donor_id'],
            'appearance_years': years_sorted,
            'appearance_count': len(years_sorted),
            'score': score,
            'is_float_buyer': is_float,
            'entity_type': entity_type,
            'donor_status': donor_status,
            'match_conf': info['best_conf'],
            'is_follow': is_follow,
        })

    results.sort(key=lambda r: (-r['score'], r['name_primary']))
    return results


# ─── Layer 4c: Analysis C — Segmentation ─────────────────────────────────────

def analyze_C(b_results: list) -> dict:
    hot, warm, cold, review = [], [], [], []
    for r in b_results:
        if r['is_float_buyer'] or r['entity_type'] == 'NGO/學生/公民社會':
            continue
        if '待確認' in r['entity_type']:
            review.append(r)
            continue
        if r['entity_type'] == '可能商業':
            has_relationship = (
                r['donor_status'] in ('級別贊助', '市集', '其他')
                or r['donor_id'] is not None  # historical sponsor, even if inactive
            )
            if has_relationship:
                hot.append(r)
            elif r['appearance_count'] >= 3:
                warm.append(r)
            else:
                cold.append(r)
    return {'hot': hot, 'warm': warm, 'cold': cold, 'review': review}


# ─── Layer 5: Excel Output ────────────────────────────────────────────────────

def write_excel(a: dict, b: list, c: dict, pending: list, output_path: Path):
    if not HAS_EXCEL:
        print('⚠ openpyxl not installed. Run: pip install openpyxl')
        _fallback_csv(a, b, c, output_path)
        return

    fill_green  = PatternFill('solid', fgColor='C6EFCE')
    fill_yellow = PatternFill('solid', fgColor='FFEB9C')
    fill_grey   = PatternFill('solid', fgColor='D9D9D9')
    fill_follow = PatternFill('solid', fgColor='FCE4D6')  # peach = Follow tag
    fill_hdr    = PatternFill('solid', fgColor='4472C4')
    font_hdr    = Font(color='FFFFFF', bold=True)

    def hdr_row(ws, values):
        ws.append(values)
        r = ws.max_row
        for col in range(1, len(values) + 1):
            ws.cell(r, col).fill = fill_hdr
            ws.cell(r, col).font = font_hdr

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # Sheet 1: 摘要
    ws = wb.create_sheet('摘要')
    ws.append(['遊行BD機會分析', ''])
    ws.append(['生成日期', str(date.today())])
    ws.append(['', ''])
    ws.append(['=== A: 歷史轉化率 ===', ''])
    ws.append(['商業車買家（去重）', a['total_float_buyers']])
    ws.append(['其中有遊行隊伍前史', a['had_team_history_before_float']])
    ws.append(['歷史轉化率', f"{a['conversion_rate_pct']}%"])
    ws.append(['', ''])
    ws.append(['=== C: BD機會清單 ===', ''])
    ws.append(['🔴 Hot (既有贊助商可升級)', len(c['hot'])])
    ws.append(['🟡 Warm (非贊助商, 3次以上遊行)', len(c['warm'])])
    ws.append(['⚪ Cold (非贊助商, 1-2次遊行)', len(c['cold'])])
    ws.append(['❓ 待確認 (實體類型待核)', len(c['review'])])
    ws.append(['', ''])
    ws.append(['=== 積分權重 ===', ''])
    for yr, w in sorted(RECENCY_WEIGHTS.items()):
        ws.append([f'{yr}年', w])
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 20

    # Sheet 2: 歷史轉化路徑
    ws = wb.create_sheet('歷史轉化路徑')
    hdr_row(ws, ['名稱', '捐款人ID', '首次商業車年份', '商業車前隊伍年份', '先有隊伍再買車?', '同年同時出現?'])
    for r in a['detail_rows']:
        row_data = [
            r['name'],
            r['donor_id'],
            r['first_float_year'],
            ', '.join(r['team_years_before_float']),
            '✓' if r['was_team_first'] else '',
            '●' if r['same_year_both'] else '',
        ]
        ws.append(row_data)
        if r['was_team_first']:
            last = ws.max_row
            ws.cell(last, 5).fill = fill_green
    ws.column_dimensions['A'].width = 40
    for col in 'BCDEF':
        ws.column_dimensions[col].width = 16

    # Sheet 3: 遊行隊伍全覽
    ws = wb.create_sheet('遊行隊伍全覽')
    years_cols = ['2019', '2020', '2022', '2023', '2024', '2025']
    hdr_row(ws, ['主要名稱', '捐款人ID', '出現次數', '積分', '實體類型', '捐款狀態', '已購商業車', 'Follow', '名稱變體'] + years_cols)
    for r in b:
        follow_tag = '★' if r.get('is_follow') else ''
        row_data = [
            r['name_primary'],
            r['donor_id'] or '',
            r['appearance_count'],
            r['score'],
            r['entity_type'],
            r['donor_status'],
            '✓' if r['is_float_buyer'] else '',
            follow_tag,
            ' | '.join(r['name_variants'][:3]),
        ] + ['●' if y in r['appearance_years'] else '' for y in years_cols]
        ws.append(row_data)
        last = ws.max_row
        fill = None
        if r.get('is_follow') and r['is_float_buyer']:
            fill = fill_follow  # peach for Follow float buyers
        elif r['is_float_buyer']:
            fill = fill_grey
        elif r['entity_type'] == '可能商業' and r['donor_status'] in ('級別贊助', '市集', '其他'):
            fill = fill_green
        elif r['entity_type'] == '可能商業' and r['appearance_count'] >= 3 and r['donor_status'] == '非捐款者':
            fill = fill_yellow
        if fill:
            for col in range(1, 10 + len(years_cols)):
                ws.cell(last, col).fill = fill
    ws.column_dimensions['A'].width = 36
    ws.column_dimensions['I'].width = 40
    ws.freeze_panes = 'A2'

    # Sheet 4: BD機會清單
    ws = wb.create_sheet('BD機會清單')
    cols = ['名稱', '捐款人ID', '出現次數', '積分', '目前捐款狀態', '出現年份', '名稱變體', '建議行動']
    sections = [
        ('🔴 HOT — 既有贊助商 × 遊行隊伍 → 升級商業車', c['hot'], fill_green, '跟進升級商業車'),
        ('🟡 WARM — 非贊助商 × 3次以上遊行 → 開發', c['warm'], fill_yellow, '主動開發接洽'),
        ('⚪ COLD — 非贊助商 × 1-2次遊行 → 名單留存', c['cold'], None, '名單留存'),
        ('❓ 待確認 — 實體類型待人工核對', c['review'], None, '確認實體類型後分類'),
    ]
    for title, rows, fill, action in sections:
        ws.append([title])
        hdr_row(ws, cols)
        for r in rows:
            ws.append([
                r['name_primary'],
                r['donor_id'] or '',
                r['appearance_count'],
                r['score'],
                r['donor_status'],
                ', '.join(r['appearance_years']),
                ' | '.join(r['name_variants'][:3]),
                action,
            ])
            last = ws.max_row
            if fill:
                for col in range(1, len(cols) + 1):
                    ws.cell(last, col).fill = fill
        ws.append([''])

    # Follow section: float buyers marked for upsell to tier sponsorship
    follow_upsell = [r for r in b if r.get('is_follow') and r['is_float_buyer']
                     and r['donor_status'] not in ('級別贊助',)]
    if follow_upsell:
        ws.append(['★ Follow — 商業車買家升級目標（已標記，尚非級別贊助）'])
        hdr_row(ws, cols)
        for r in follow_upsell:
            ws.append([
                r['name_primary'],
                r['donor_id'] or '',
                r['appearance_count'],
                r['score'],
                r['donor_status'],
                ', '.join(r['appearance_years']),
                ' | '.join(r['name_variants'][:3]),
                '跟進升級級別贊助',
            ])
            last = ws.max_row
            for col in range(1, len(cols) + 1):
                ws.cell(last, col).fill = fill_follow
        ws.append([''])
    ws.column_dimensions['A'].width = 36
    ws.column_dimensions['G'].width = 40
    ws.freeze_panes = 'A2'

    # Sheet 5: 待人工確認 (commercial-only; NGO/civic filtered out)
    commercial_raws: set = set()
    for r in b:
        if r['entity_type'] == '可能商業' and not r['donor_id']:
            for v in r['name_variants']:
                commercial_raws.add(v)
    raw_to_entity = {}
    for r in b:
        for v in r['name_variants']:
            raw_to_entity[v] = r

    ws = wb.create_sheet('待人工確認')
    ws.append(['=== A: 商業類名稱比對未確認（NGO/學生/公民類已過濾）==='])
    hdr_row(ws, ['原始名稱', '正規化後', '出現次數', '出現年份', '積分', '備注'])
    written: set = set()
    for p in pending:
        raw = p['raw_name']
        if raw in written or raw not in commercial_raws:
            continue
        written.add(raw)
        er = raw_to_entity.get(raw, {})
        ws.append([
            raw,
            p['normalized'],
            er.get('appearance_count', ''),
            ', '.join(er.get('appearance_years', [])),
            er.get('score', ''),
            '請確認是否在捐款人名單中',
        ])
    ws.append([''])
    ws.append(['=== B: 實體類型待確認 ==='])
    hdr_row(ws, ['名稱', '捐款人ID', '出現年份', '目前分類', '建議分類'])
    for r in b:
        if '待確認' in r['entity_type']:
            ws.append([
                r['name_primary'],
                r['donor_id'] or '',
                ', '.join(r['appearance_years']),
                r['entity_type'],
                '',
            ])
    ws.column_dimensions['A'].width = 42
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 8
    ws.column_dimensions['D'].width = 28
    ws.column_dimensions['E'].width = 10

    wb.save(output_path)
    print(f'Excel saved: {output_path}')


def _fallback_csv(a, b, c, output_path: Path):
    import csv
    csv_path = output_path.with_suffix('.csv')
    with open(csv_path, 'w', newline='', encoding='utf-8-sig') as f:
        writer = csv.writer(f)
        writer.writerow(['名稱', '捐款人ID', '出現次數', '積分', '實體類型', '捐款狀態', '已購商業車', '出現年份'])
        for r in b:
            writer.writerow([
                r['name_primary'], r['donor_id'] or '', r['appearance_count'],
                r['score'], r['entity_type'], r['donor_status'],
                '✓' if r['is_float_buyer'] else '',
                ', '.join(r['appearance_years']),
            ])
    print(f'CSV saved (no openpyxl): {csv_path}')


# ─── Console Summary ──────────────────────────────────────────────────────────

def print_summary(a: dict, c: dict, b: list, pending: list):
    print('\n' + '=' * 55)
    print('遊行BD機會分析')
    print('=' * 55)
    print(f'\n[A] 歷史轉化率')
    print(f'  商業車買家（去重）: {a["total_float_buyers"]}')
    print(f'  其中有遊行隊伍前史: {a["had_team_history_before_float"]} ({a["conversion_rate_pct"]}%)')
    print(f'\n[C] BD機會清單')
    print(f'  🔴 Hot  (既有贊助商可升級): {len(c["hot"])}')
    print(f'  🟡 Warm (非贊助商, 3次+):   {len(c["warm"])}')
    print(f'  ⚪ Cold (非贊助商, 1-2次):  {len(c["cold"])}')
    print(f'  ❓ 待確認:                  {len(c["review"])}')
    if c['hot']:
        print('\n  Top Hot Prospects:')
        for r in c['hot'][:5]:
            print(f'    {(r["donor_id"] or "???"):8s}  {r["name_primary"][:32]:32s}  {r["appearance_count"]}次  {r["donor_status"]}')
    if c['warm']:
        print('\n  Top Warm Prospects:')
        for r in c['warm'][:8]:
            print(f'    {(r["donor_id"] or "???"):8s}  {r["name_primary"][:32]:32s}  {r["appearance_count"]}次  {"、".join(r["appearance_years"])}')
    print(f'\n  待人工確認名稱數: {len(pending)}')
    # Show entity breakdown
    types = {}
    for r in b:
        types[r['entity_type']] = types.get(r['entity_type'], 0) + 1
    print('\n  實體類型分布:')
    for k, v in sorted(types.items(), key=lambda x: -x[1]):
        print(f'    {k:30s}: {v}')


# ─── Main ─────────────────────────────────────────────────────────────────────

def main():
    base = Path(__file__).parent

    print('Parsing parade data from MD file...')
    md_data = parse_md(base / 'taiwan-pride-parade-teams-2019-2025.md')
    for year in sorted(md_data):
        d = md_data[year]
        f = len(d['commercial_floats'])
        t = len(d['parade_teams'])
        a2021 = len(d['2021_all']) if d['2021_all'] else 0
        print(f'  {year}: {f} 商業車, {t} 隊伍, {a2021} 全體(2021)')

    print('\nLoading donor database...')
    db = load_donors(base)

    matcher = NameMatcher(db)

    print('\nRunning Analysis A (historical conversion)...')
    a = analyze_A(md_data, db, matcher)

    print('Running Analysis B (prospect scoring)...')
    b = analyze_B(md_data, db, matcher)

    print('Running Analysis C (segmentation)...')
    c = analyze_C(b)

    output_path = base / f'遊行BD機會分析_{date.today():%Y%m%d}.xlsx'
    write_excel(a, b, c, matcher.pending, output_path)

    print_summary(a, c, b, matcher.pending)


if __name__ == '__main__':
    main()
